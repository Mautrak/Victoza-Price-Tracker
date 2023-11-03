import re
import openpyxl
from openpyxl.styles import NamedStyle
import requests
import time
import threading
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options as FirefoxOptions
import winsound

# Constants
URL1 = "https://liki24.com/uk/p/viktoza-r-r-d-6-mgml-kartridzh-vlozh-v-shpric-ruchku-3-ml-2-novo-nordisk/"
URL2 = "https://tabletki.ua/Виктоза/25550/"
XPATH1 = "/html/body/div[3]/div[1]/div[5]/div[2]/div[2]/div[3]/div[2]/div[1]/div[3]/span"
XPATH2 = "/html/body/div[2]/main/div[1]/div[1]/div[1]/article/div/div/section[1]/div/div[2]/div[1]/div[2]/div[1]/div[1]/span"
EXCEL_FILE = "prices.xlsx"
ALARM_FREQUENCY = 2500
ALARM_DURATION = 1000
COST_COLUMN = 1
DATE_COLUMN = 2
SITE_COLUMN = 3

stop_event = threading.Event()

def is_connected():
    try:
        requests.get("https://www.google.com", timeout=5)
        return True
    except requests.ConnectionError:
        return False

def wait_until_next_run():
    now = datetime.now()
    next_run = datetime(now.year, now.month, now.day, 11, 0) + timedelta(days=1)
    while datetime.now() < next_run:
        if stop_event.is_set():
            return
        time.sleep(60)

def get_price(driver, url, xpath):
    driver.get(url)
    price_text = driver.find_element(By.XPATH, xpath).text
    price = re.search(r'[\d\s]+[\.,]?\d*', price_text).group()
    price = price.replace(' ', '').replace(',', '.')
    return float(price)

def main():
    currency_style = NamedStyle(name='currency_style', number_format='#,##0.00 [$₴-422]')

    while not stop_event.is_set():
        today = datetime.now().date()
        try:
            workbook = openpyxl.load_workbook(EXCEL_FILE)
            sheet = workbook.active
            for col_cells in sheet.iter_cols(min_col=COST_COLUMN, max_col=COST_COLUMN):
                for cell in col_cells[1:]:  # Skipping the header
                    cell.style = currency_style
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["Cost of Victoza", "Date", "Site", "Difference (UAH)", "Difference (%)"])
            for col_cells in sheet.iter_cols(min_col=COST_COLUMN, max_col=COST_COLUMN):
                for cell in col_cells[1:]:
                    cell.style = currency_style
            workbook.save(EXCEL_FILE)

        while not is_connected():
            print("No internet connection. Waiting...")
            time.sleep(60)

        options = FirefoxOptions()
        driver = webdriver.Firefox(options=options)

        print("Connecting to URL1...")
        try:
            price1 = get_price(driver, URL1, XPATH1)
            print(f"Successfully reached URL1. Price found: {price1}")
        except Exception as e:
            print(f"Failed to get price from URL1: {e}")
            price1 = float('inf')

        print("Connecting to URL2...")
        try:
            price2 = get_price(driver, URL2, XPATH2)
            print(f"Successfully reached URL2. Price found: {price2}")
        except Exception as e:
            print(f"Failed to get price from URL2: {e}")
            price2 = float('inf')

        final_price = min(price1, price2)
        site_used = "Liki24" if final_price == price1 else "Tabletki"

        diff_uah = 0
        diff_percentage = 0
        last_row = sheet.max_row

        if last_row > 1:
            yesterday_price = sheet.cell(row=last_row, column=COST_COLUMN).value
            diff_uah = final_price - yesterday_price
            diff_percentage = (diff_uah / yesterday_price) * 100

        original_price = 1000.00
        if final_price <= original_price * 0.7:
            winsound.Beep(ALARM_FREQUENCY, ALARM_DURATION)

        print("Updating the Excel sheet with the price.")
        sheet.append([final_price, today.strftime('%Y-%m-%d'), site_used, diff_uah, diff_percentage])
        workbook.save(EXCEL_FILE)

        driver.quit()
        print("Done for today. Waiting until tomorrow.")
        wait_until_next_run()

if __name__ == "__main__":
    main()